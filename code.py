import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font, Border, Side, Alignment

st.set_page_config(page_title="CKD Position Validator", layout="wide")
st.markdown("# 📺 Vérificateur de Positions CKD")
st.markdown("Vérifie que le nombre de positions correspond à la quantité pour les composants CKD uniquement")

# Upload du fichier BOM
old_file = st.file_uploader("📂 Upload votre fichier BOM", type=["xlsx"])

def extract_ckd_components(df):
    """
    Extrait uniquement les composants CKD
    CKD: de "ASS'Y - MAIN BOARD（CKD）" jusqu'à "BARCODE LABEL"
    """
    start_idx = None
    for idx, desc in enumerate(df['Description']):
        if 'ASS\'Y - MAIN BOARD（CKD）' in str(desc).upper() or 'ASSY - MAIN BOARD（CKD）' in str(desc).upper():
            start_idx = idx
            break
    
    end_idx = None
    for idx, desc in enumerate(df['Description']):
        if 'BARCODE LABEL' in str(desc).upper():
            end_idx = idx
            break
    
    if start_idx is not None and end_idx is not None:
        return df.iloc[start_idx:end_idx+1].copy()
    elif start_idx is not None:
        return df.iloc[start_idx:].copy()
    else:
        return pd.DataFrame()

def safe_join(x):
    """Convertit une liste en chaîne de caractères"""
    if not isinstance(x, list):
        return str(x) if pd.notna(x) else ""
    return ", ".join(str(i) for i in x if pd.notna(i))

def extract_positions(bom_text):
    """
    Extrait les positions individuelles depuis BOM text
    """
    positions = []
    if bom_text and str(bom_text) != "nan":
        # Convertir en string et nettoyer
        bom_text_str = str(bom_text)
        # Séparer par virgule
        raw_positions = bom_text_str.split(',')
        for pos in raw_positions:
            pos = pos.strip()
            # Nettoyer les crochets, guillemets, etc.
            pos = pos.replace('[', '').replace(']', '').replace("'", "").replace('"', "").strip()
            if pos and pos != "nan":
                positions.append(pos)
    return positions

def is_non_component(description):
    """
    Vérifie si la ligne n'est pas un composant (maison mère, PCB, etc.)
    """
    non_components = [
        'ASS\'Y - MAIN BOARD（CKD）',
        'ASSY - MAIN BOARD（CKD）',
        'ASS\'Y - MAIN BOARD',
        'ASSY - MAIN BOARD',
        'ASS\'Y - MAIN BOARD（SMT）',
        'ASSY - MAIN BOARD（SMT）',
        'PCB',
        'THERMAL CONDUCTIVE',
        'BARCODE LABEL'
    ]
    
    desc_upper = str(description).upper()
    for non_comp in non_components:
        if non_comp.upper() in desc_upper:
            return True
    return False

def validate_ckd_positions(df):
    """
    Vérifie la cohérence entre les positions et la quantité pour CKD
    """
    results = []
    
    for idx, row in df.iterrows():
        # Récupérer les données
        pn = row.get("PN", "")
        description = row.get("Description", "")
        bom_text = row.get("BOM text", "")
        qty = row.get("bom_qty", 0)
        
        # Vérifier si c'est un non-composant
        is_non_comp = is_non_component(description)
        
        # Nettoyer la quantité
        try:
            if isinstance(qty, str):
                qty = qty.replace(",", ".")
            qty = float(qty) if qty else 0
        except:
            qty = 0
        
        # Extraire les positions
        positions = extract_positions(bom_text)
        nb_positions = len(positions)
        positions_str = safe_join(positions)
        
        # Déterminer le résultat
        if is_non_comp:
            result = "📌 NO NEED / NOT APPLICABLE"
            color = "#D9E1F2"  # Bleu clair
        elif nb_positions == 0 and qty == 0:
            result = "✅ VIDE"
            color = "#E2EFDA"  # Vert très clair
        elif nb_positions == 0 and qty > 0:
            result = "❌ ERREUR - Aucune position"
            color = "#FFC7CE"  # Rouge
        elif nb_positions == qty:
            result = "✅ CONFORME"
            color = "#C6EFCE"  # Vert
        elif nb_positions < qty:
            result = f"⚠️ MANQUE - {int(qty - nb_positions)} position(s) manquante(s)"
            color = "#FFEB9C"  # Jaune
        else:  # nb_positions > qty
            result = f"⚠️ TROP - {int(nb_positions - qty)} position(s) en excès"
            color = "#FFEB9C"  # Jaune
        
        results.append({
            "PN": pn,
            "Description": description,
            "QTY": int(qty) if qty == int(qty) else qty,
            "Position": positions_str,
            "QTY Calculated": nb_positions,
            "Result": result,
            "_color": color,
            "_is_non_component": is_non_comp
        })
    
    return pd.DataFrame(results)

def export_to_colored_excel(df, filename):
    """
    Exporte le DataFrame vers Excel avec couleurs
    """
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Verification_CKD', index=False)
        
        # Appliquer les couleurs
        workbook = writer.book
        worksheet = writer.sheets['Verification_CKD']
        
        # Définir les couleurs pour chaque statut
        color_map = {
            "📌 NO NEED / NOT APPLICABLE": "D9E1F2",  # Bleu clair
            "✅ CONFORME": "C6EFCE",  # Vert
            "❌ ERREUR - Aucune position": "FFC7CE",  # Rouge
            "✅ VIDE": "E2EFDA",  # Vert très clair
        }
        
        # Pour les statuts avec "MANQUE" ou "TROP"
        warning_color = "FFEB9C"  # Jaune
        
        # Parcourir les lignes
        for row in range(2, worksheet.max_row + 1):
            status_cell = worksheet.cell(row=row, column=6)  # Colonne Result (6ème colonne)
            status = status_cell.value
            
            # Déterminer la couleur
            if status in color_map:
                color = color_map[status]
            elif "MANQUE" in str(status) or "TROP" in str(status):
                color = warning_color
            else:
                color = "FFFFFF"  # Blanc par défaut
            
            # Appliquer la couleur à toute la ligne
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for col in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row, column=col).fill = fill
            
            # Ajouter des bordures
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            for col in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row, column=col).border = thin_border
        
        # Mettre en gras l'en-tête
        for col in range(1, worksheet.max_column + 1):
            header_cell = worksheet.cell(row=1, column=col)
            header_cell.font = Font(bold=True)
            header_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_cell.font = Font(bold=True, color="FFFFFF")
        
        # Ajuster les largeurs des colonnes
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    return output

if old_file:
    # Lire le fichier Excel
    try:
        df = pd.read_excel(old_file)
        
        # Nettoyer les noms de colonnes
        df.columns = df.columns.str.strip()
        
        # Afficher un aperçu du fichier
        with st.expander("📋 Aperçu du fichier chargé"):
            st.dataframe(df.head(10), use_container_width=True)
        
        # Vérifier que les colonnes nécessaires existent
        required_cols = ["bom_qty", "BOM text"]
        missing_cols = [col for col in required_cols if col not in df.columns]
        
        if missing_cols:
            st.error(f"❌ Colonnes manquantes dans le fichier: {', '.join(missing_cols)}")
            st.info(f"Les colonnes disponibles sont: {', '.join(df.columns)}")
        else:
            # Extraire uniquement la partie CKD
            with st.spinner("Extraction des composants CKD..."):
                ckd_df = extract_ckd_components(df)
            
            if ckd_df.empty:
                st.warning("⚠️ Aucun composant CKD trouvé dans le fichier")
                st.info("Assurez-vous que votre fichier contient la ligne 'ASS'Y - MAIN BOARD（CKD）' et 'BARCODE LABEL'")
            else:
                st.success(f"✅ {len(ckd_df)} composants CKD extraits")
                
                # Lancer la validation
                if st.button("🔍 Vérifier les positions CKD", use_container_width=True):
                    with st.spinner("Vérification en cours..."):
                        results_df = validate_ckd_positions(ckd_df)
                        
                        # Statistiques
                        st.subheader("📊 Résumé de la vérification CKD")
                        
                        col1, col2, col3, col4, col5 = st.columns(5)
                        total = len(results_df)
                        conformes = len(results_df[results_df["Result"] == "✅ CONFORME"])
                        erreurs = len(results_df[results_df["Result"].str.contains("ERREUR", na=False)])
                        manques = len(results_df[results_df["Result"].str.contains("MANQUE", na=False)])
                        trop = len(results_df[results_df["Result"].str.contains("TROP", na=False)])
                        no_need = len(results_df[results_df["Result"] == "📌 NO NEED / NOT APPLICABLE"])
                        
                        with col1:
                            st.metric("📊 Total", total)
                        with col2:
                            st.metric("✅ Conformes", conformes)
                        with col3:
                            st.metric("❌ Erreurs", erreurs, delta="À corriger" if erreurs > 0 else None, delta_color="inverse")
                        with col4:
                            st.metric("⚠️ Manque/Trop", manques + trop)
                        with col5:
                            st.metric("📌 Non applicable", no_need)
                        
                        # Afficher le tableau des résultats
                        st.subheader("🔍 Détail de la vérification CKD")
                        
                        # Sélectionner uniquement les colonnes souhaitées
                        display_cols = ["PN", "Description", "QTY", "Position", "QTY Calculated", "Result"]
                        display_df = results_df[display_cols].copy()
                        
                        # CORRECTION : Utiliser map au lieu de applymap
                        def color_result(val):
                            if "CONFORME" in str(val):
                                return 'background-color: #C6EFCE'
                            elif "ERREUR" in str(val):
                                return 'background-color: #FFC7CE'
                            elif "MANQUE" in str(val) or "TROP" in str(val):
                                return 'background-color: #FFEB9C'
                            elif "NO NEED" in str(val):
                                return 'background-color: #D9E1F2'
                            return ''
                        
                        # Appliquer le style avec map
                        styled_df = display_df.style.map(color_result, subset=['Result'])
                        st.dataframe(styled_df, use_container_width=True)
                        
                        # Option pour filtrer les problèmes uniquement (CORRIGÉ)
                        show_problems_only = st.checkbox("⚠️ Afficher uniquement les lignes non conformes", value=False)
                        
                        if show_problems_only:
                            # Filtrer les lignes non conformes (erreurs, manques, trop)
                            problem_df = display_df[
                                (display_df["Result"].str.contains("ERREUR", na=False)) |
                                (display_df["Result"].str.contains("MANQUE", na=False)) |
                                (display_df["Result"].str.contains("TROP", na=False))
                            ].copy()
                            
                            if len(problem_df) > 0:
                                st.warning(f"⚠️ {len(problem_df)} ligne(s) avec problèmes")
                                styled_problem_df = problem_df.style.map(color_result, subset=['Result'])
                                st.dataframe(styled_problem_df, use_container_width=True)
                            else:
                                st.success("✅ Aucune ligne non conforme trouvée !")
                        else:
                            # Afficher le tableau complet avec style
                            st.dataframe(styled_df, use_container_width=True)
                        
                        # Export Excel avec couleurs
                        colored_excel = export_to_colored_excel(results_df[display_cols], "verification_positions_CKD.xlsx")
                        
                        st.download_button(
                            "📥 Télécharger le rapport Excel (coloré)",
                            colored_excel,
                            "verification_positions_CKD.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        
                        # Afficher les exemples expliqués
                        with st.expander("ℹ️ Comprendre la vérification CKD"):
                            st.markdown("""
                            ### Comment fonctionne la vérification ?
                            
                            La vérification compare le **nombre de positions** dans la colonne `BOM text` avec la **quantité** (`bom_qty`) pour les composants CKD uniquement.
                            
                            **La partie SKD n'est pas vérifiée car elle n'a pas de positions.**
                            
                            **Cas particuliers (NO NEED / NOT APPLICABLE) :**
                            - ASS'Y - MAIN BOARD（CKD）
                            - ASS'Y - MAIN BOARD
                            - ASS'Y - MAIN BOARD（SMT）
                            - PCB
                            - Thermal conductive
                            - BARCODE LABEL
                            
                            Ces éléments sont des supports ou des repères, pas des composants à positionner.
                            
                            #### Exemples:
                            
                            | BOM text | Quantity | QTY Calculated | Résultat |
                            |----------|----------|----------------|----------|
                            | DT9 | 1 | 1 | ✅ CONFORME |
                            | DT7, DT8, DT10 | 3 | 3 | ✅ CONFORME |
                            | DT7, DT8, DT10 | 2 | 3 | ⚠️ TROP - 1 position(s) en excès |
                            | DT7, DT8 | 3 | 2 | ⚠️ MANQUE - 1 position(s) manquante(s) |
                            | (vide) | 1 | 0 | ❌ ERREUR - Aucune position |
                            | ASS'Y - MAIN BOARD | 1 | 0 | 📌 NO NEED / NOT APPLICABLE |
                            
                            **Format accepté pour les positions :**
                            - Positions séparées par des virgules : `DT1, DT2, DT3`
                            - Avec ou sans espaces
                            - Avec ou sans crochets `[DT1, DT2]`
                            """)
                        
            # Afficher la partie SKD (non vérifiée)
            with st.expander("ℹ️ Note sur la partie SKD"):
                st.info("""
                **La partie SKD n'est pas vérifiée** car les composants SKD n'ont pas de positions (colonne `BOM text` vide).
                
                La vérification se concentre uniquement sur les composants CKD qui se trouvent entre:
                - **Début:** `ASS'Y - MAIN BOARD（CKD）`
                - **Fin:** `BARCODE LABEL`
                """)
                
                # Compter les composants SKD
                start_idx = None
                for idx, desc in enumerate(df['Description']):
                    if 'ASS\'Y - MAIN BOARD（CKD）' in str(desc).upper() or 'ASSY - MAIN BOARD（CKD）' in str(desc).upper():
                        start_idx = idx
                        break
                
                end_idx = None
                for idx, desc in enumerate(df['Description']):
                    if 'BARCODE LABEL' in str(desc).upper():
                        end_idx = idx
                        break
                
                if start_idx is not None:
                    if end_idx is not None:
                        skd_df = pd.concat([df.iloc[:start_idx], df.iloc[end_idx+1:]], ignore_index=True)
                    else:
                        skd_df = df.iloc[:start_idx]
                    st.write(f"📦 Nombre de composants SKD: {len(skd_df)}")
        
    except Exception as e:
        st.error(f"❌ Erreur lors de la lecture du fichier: {str(e)}")
else:
    st.info("👆 Veuillez uploader un fichier Excel pour commencer la vérification")
    
    # Afficher un exemple
    with st.expander("📖 Format attendu du fichier"):
        st.markdown("""
        ### Colonnes requises dans votre fichier Excel:
        
        | PN | Description | bom_qty | BOM text |
        |-----|-------------|---------|----------|
        | R001 | Résistance | 1 | R12 |
        | C002 | Condensateur | 3 | C1, C2, C3 |
        | IC003 | Circuit intégré | 2 | U1, U2 |
        
        ### Structure attendue pour CKD:
        
        Le fichier doit contenir les lignes de délimitation:
        1. **Ligne de début CKD:** `ASS'Y - MAIN BOARD（CKD）`
        2. **Ligne de fin CKD:** `BARCODE LABEL`
        
        ### Ce que fait le vérificateur:
        
        1. Extrait uniquement les composants **entre** ces deux lignes (partie CKD)
        2. Compte le nombre de positions dans la colonne `BOM text`
        3. Compare avec la quantité (`bom_qty`)
        4. Affiche le résultat dans les colonnes:
           - `QTY Calculated` = nombre de positions trouvées
           - `Result` = statut de la vérification
        
        **La partie SKD (hors délimiteurs) n'est pas vérifiée car elle n'a pas de positions.**
        """)
